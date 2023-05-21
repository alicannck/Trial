from selenium import webdriver    
from webdriver_manager.chrome import ChromeDriverManager 
from selenium.webdriver.common.by import By 
from time import sleep  
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import Globalconstants


class Test_saucedemo:

    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(Globalconstants.URL)
        self.driver.maximize_window()

        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    
    def teardown_method(self):
        self.driver.quit()

    def Web_Wait(self,locator):
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located(locator))
    

    def test_Practice1(self):

        self.Web_Wait(Globalconstants.NAME_ID)
        name = self.driver.find_element(By.ID,"user-name")
        name.send_keys("")

        self.Web_Wait(Globalconstants.PASSWORD_ID)
        password = self.driver.find_element(By.ID,"password")
        password.send_keys("")

        self.Web_Wait(Globalconstants.BUTTON_ID)
        button = self.driver.find_element(By.ID,"login-button")
        button.click()

        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]")
        assert errorMessage.text == (Globalconstants.errorMessage1)

        self.driver.save_screenshot(f"{self.folderPath}/test_Practice_1.png")



    def test_Practice2(self):

        self.Web_Wait(Globalconstants.NAME_ID)
        name = self.driver.find_element(By.ID,"user-name")
        name.send_keys("standard_user")

        self.Web_Wait(Globalconstants.PASSWORD_ID)
        password = self.driver.find_element(By.ID,"password")
        password.send_keys("")

        self.Web_Wait(Globalconstants.BUTTON_ID)
        button = self.driver.find_element(By.ID,"login-button")
        button.click()

        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]")
        assert errorMessage.text == (Globalconstants.errorMessage2)

        self.driver.save_screenshot(f"{self.folderPath}/test_Practice_2.png")
    

    
    def test_Practice3(self):

        self.Web_Wait(Globalconstants.NAME_ID)
        name = self.driver.find_element(By.ID,"user-name")
        name.send_keys("locked_out_user")

        self.Web_Wait(Globalconstants.PASSWORD_ID)
        password = self.driver.find_element(By.ID,"password")
        password.send_keys("secret_sauce")

        self.Web_Wait(Globalconstants.BUTTON_ID)
        button = self.driver.find_element(By.ID,"login-button")
        button.click()

        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]")
        assert errorMessage.text == (Globalconstants.errorMessage3)
    
        self.driver.save_screenshot(f"{self.folderPath}/test_Practice_3.png")



    def test_Practice4(self):
        
        self.Web_Wait(Globalconstants.NAME_ID)
        name = self.driver.find_element(By.ID,"user-name")
        name.send_keys("standard_user")

        self.Web_Wait(Globalconstants.PASSWORD_ID)
        password = self.driver.find_element(By.ID,"password")
        password.send_keys("secret_sauce")

        self.Web_Wait(Globalconstants.BUTTON_ID)
        button = self.driver.find_element(By.ID,"login-button")
        button.click()

        self.driver.get("https://www.saucedemo.com/inventory.html")

        self.driver.save_screenshot(f"{self.folderPath}/test_Practice_4.png")



    @pytest.mark.parametrize("username,password",Globalconstants.LIST)
    def test_Multi_Pratice(self,username,password):

        self.Web_Wait(Globalconstants.NAME_ID)
        name = self.driver.find_element(By.ID,"user-name")
        name.send_keys(username)

        self.Web_Wait(Globalconstants.PASSWORD_ID)
        passw = self.driver.find_element(By.ID,"password")
        passw.send_keys(password)

        self.Web_Wait(Globalconstants.BUTTON_ID)
        button = self.driver.find_element(By.ID,"login-button")
        button.click()

        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]")
        assert errorMessage.text == (Globalconstants.errorMessage4)

        self.driver.save_screenshot(f"{self.folderPath}/test_Multi_Practice_{username},{password}.png")



    def getData():
        Excelfile = openpyxl.load_workbook("data/into.xlsx")
        page = Excelfile["Sayfa1"]

        totalRow = page.max_row 
        data=[]
        for i in range(2,totalRow+1): 
            username = page.cell(i,1).value 
            password = page.cell(i,2).value 
            tupleData = (username,password)
            data.append(tupleData)
       
        return data


    @pytest.mark.parametrize("username,password",getData())
    def test_Excel_Practice(self,username,password):

        self.Web_Wait(Globalconstants.NAME_ID)
        name = self.driver.find_element(By.ID,"user-name")
        name.send_keys(username)

        self.Web_Wait(Globalconstants.PASSWORD_ID)
        passw = self.driver.find_element(By.ID,"password")
        passw.send_keys(password)

        self.Web_Wait(Globalconstants.BUTTON_ID)
        button = self.driver.find_element(By.ID,"login-button")
        button.click()

        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]")
        assert errorMessage.text == (Globalconstants.errorMessage4)
        
        self.driver.save_screenshot(f"{self.folderPath}/test_Excel_Practice_{username},{password}.png")
    
    